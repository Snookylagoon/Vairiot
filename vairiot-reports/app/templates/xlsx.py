"""
XLSX report generator — openpyxl with full Vairiot livery.

Layout (A4):
  Row 1-2   : Merged header band (gradient fill → solid violet, logo text + title)
  Row 3     : Metadata row (tenant, filters, date) on v-wash background
  Row 4     : Summary metrics row (if report has summary_fields)
  Row 5     : Blank spacer
  Row 6     : Column headers (v-violet background, white text)
  Row 7+    : Data rows (alternating white / v-wash)
  Last row  : Totals row (v-violet background, white text) if show_totals
  Footer    : Print footer with company info + page numbers
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintPageSetup

from app.config import (
    BRAND, ColType, ReportDef,
    A4_WIDTH_MM, A4_HEIGHT_MM,
    MARGIN_TOP_MM, MARGIN_BOTTOM_MM, MARGIN_LEFT_MM, MARGIN_RIGHT_MM,
)
from app.models import ReportRequest
from app.templates.brand import format_value, hex_to_rgb


# ── Colour fills ──────────────────────────────────────────────────────────────

_VIOLET_FILL = PatternFill(start_color="615AA0", end_color="615AA0", fill_type="solid")
_WASH_FILL   = PatternFill(start_color="F8F0FA", end_color="F8F0FA", fill_type="solid")
_WHITE_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_PINK_FILL   = PatternFill(start_color="FF0DCC", end_color="FF0DCC", fill_type="solid")
_MAUVE_FILL  = PatternFill(start_color="A05B97", end_color="A05B97", fill_type="solid")

_THIN_BORDER_SIDE = Side(style="thin", color="D9D9D9")
_THIN_BORDER = Border(
    bottom=_THIN_BORDER_SIDE,
)


# ── Fonts ─────────────────────────────────────────────────────────────────────

_FONT_HEADER_BRAND  = Font(name="Montserrat", size=16, bold=True, color="FFFFFF")
_FONT_HEADER_SUB    = Font(name="Montserrat", size=8, color="FFFFFF")
_FONT_TITLE         = Font(name="Montserrat", size=11, bold=True, color="FFFFFF")
_FONT_META_LABEL    = Font(name="Montserrat", size=8, color="615AA0")
_FONT_META_VALUE    = Font(name="Montserrat", size=8, bold=True, color="2B3132")
_FONT_SUMMARY_VALUE = Font(name="Montserrat", size=12, bold=True, color="615AA0")
_FONT_SUMMARY_LABEL = Font(name="Montserrat", size=7, color="888888")
_FONT_COL_HEADER    = Font(name="Montserrat", size=8, bold=True, color="FFFFFF")
_FONT_DATA          = Font(name="Montserrat", size=8, color="2B3132")
_FONT_DATA_MONO     = Font(name="IBM Plex Mono", size=8, color="2B3132")
_FONT_TOTALS        = Font(name="Montserrat", size=8, bold=True, color="FFFFFF")
_FONT_TOTALS_MONO   = Font(name="IBM Plex Mono", size=8, bold=True, color="FFFFFF")
_FONT_FOOTER        = Font(name="Montserrat", size=7, color="888888")


# ── Alignment ─────────────────────────────────────────────────────────────────

_ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _col_alignment(col_type: str) -> Alignment:
    if col_type in (ColType.CURRENCY, ColType.NUMBER, ColType.INTEGER, ColType.PERCENT):
        return _ALIGN_RIGHT
    if col_type == ColType.DATE:
        return _ALIGN_CENTER
    return _ALIGN_LEFT


def _col_font(col_type: str, is_totals: bool = False) -> Font:
    if is_totals:
        if col_type in (ColType.CURRENCY, ColType.NUMBER, ColType.INTEGER):
            return _FONT_TOTALS_MONO
        return _FONT_TOTALS
    if col_type in (ColType.CURRENCY, ColType.NUMBER, ColType.INTEGER):
        return _FONT_DATA_MONO
    return _FONT_DATA


def _mm_to_excel_col_width(mm: float) -> float:
    """Approximate conversion: 1 Excel width unit ≈ 2.3 mm at default zoom."""
    return mm / 2.3


def _mm_to_excel_row_height(mm: float) -> float:
    """Approximate conversion: 1 point ≈ 0.35 mm."""
    return mm / 0.35


def generate_xlsx(report_def: ReportDef, req: ReportRequest) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = report_def.title[:31]  # Excel tab name limit

    num_cols = len(report_def.columns)
    last_col_letter = get_column_letter(num_cols)

    # ── Page setup ────────────────────────────────────────────────────────
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = (
        "landscape" if report_def.is_landscape else "portrait"
    )

    # Margins (openpyxl uses inches)
    ws.page_margins.left   = MARGIN_LEFT_MM / 25.4
    ws.page_margins.right  = MARGIN_RIGHT_MM / 25.4
    ws.page_margins.top    = MARGIN_TOP_MM / 25.4
    ws.page_margins.bottom = MARGIN_BOTTOM_MM / 25.4
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    # Print area and repeating headers will be set after we know the row count
    ws.sheet_properties.pageSetUpPr.fitToPage = False

    # ── Column widths ─────────────────────────────────────────────────────
    page_w_mm = report_def.page_width_mm - MARGIN_LEFT_MM - MARGIN_RIGHT_MM
    col_widths_pt = report_def.resolve_column_widths_pt()
    # Convert points → mm → Excel width units
    for i, pt_w in enumerate(col_widths_pt):
        mm_w = pt_w * 25.4 / 72.0
        excel_w = _mm_to_excel_col_width(mm_w)
        ws.column_dimensions[get_column_letter(i + 1)].width = max(excel_w, 6)

    current_row = 1

    # ── Row 1-2: Header band ─────────────────────────────────────────────
    # Row 1: gradient accent strip (3 cells: pink, mauve, violet across columns)
    ws.row_dimensions[current_row].height = 4
    accent_split = max(1, num_cols // 3)
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=current_row, column=c)
        if c <= accent_split:
            cell.fill = _PINK_FILL
        elif c <= accent_split * 2:
            cell.fill = _MAUVE_FILL
        else:
            cell.fill = _VIOLET_FILL
    current_row += 1

    # Row 2-3: Main header band (violet background)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=max(1, num_cols // 2))
    brand_cell = ws.cell(row=current_row, column=1)
    brand_cell.value = "VAIRIOT"
    brand_cell.font = _FONT_HEADER_BRAND
    brand_cell.fill = _VIOLET_FILL
    brand_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Report title on right side
    title_start = max(2, num_cols // 2 + 1)
    ws.merge_cells(start_row=current_row, start_column=title_start,
                   end_row=current_row, end_column=num_cols)
    title_cell = ws.cell(row=current_row, column=title_start)
    title_cell.value = report_def.title
    title_cell.font = _FONT_TITLE
    title_cell.fill = _VIOLET_FILL
    title_cell.alignment = Alignment(horizontal="right", vertical="center")

    # Fill remaining cells in header row
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=current_row, column=c)
        cell.fill = _VIOLET_FILL
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Row 3: Subtitle / tag line
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=max(1, num_cols // 2))
    sub_cell = ws.cell(row=current_row, column=1)
    sub_cell.value = "ASSET INTELLIGENCE"
    sub_cell.font = _FONT_HEADER_SUB
    sub_cell.fill = _VIOLET_FILL
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Generated date on right
    gen_date_start = max(2, num_cols // 2 + 1)
    ws.merge_cells(start_row=current_row, start_column=gen_date_start,
                   end_row=current_row, end_column=num_cols)
    date_cell = ws.cell(row=current_row, column=gen_date_start)
    date_cell.value = f"Generated: {date.today().strftime('%d %b %Y')}"
    date_cell.font = _FONT_HEADER_SUB
    date_cell.fill = _VIOLET_FILL
    date_cell.alignment = Alignment(horizontal="right", vertical="center")

    for c in range(1, num_cols + 1):
        ws.cell(row=current_row, column=c).fill = _VIOLET_FILL
    ws.row_dimensions[current_row].height = 16
    current_row += 1

    # ── Row 4: Metadata row ──────────────────────────────────────────────
    meta_items = []
    if req.tenant_name:
        meta_items.append(("Tenant", req.tenant_name))
    for k, v in req.filters.items():
        label = k.replace("_", " ").replace("Id", "").title()
        meta_items.append((label, str(v)))
    meta_items.append(("Total rows", str(len(req.rows))))

    ws.row_dimensions[current_row].height = 18
    col_idx = 1
    cols_per_meta = max(1, num_cols // max(len(meta_items), 1))
    for label, value in meta_items:
        end_col = min(col_idx + cols_per_meta - 1, num_cols)
        if cols_per_meta > 1:
            ws.merge_cells(start_row=current_row, start_column=col_idx,
                           end_row=current_row, end_column=end_col)
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = f"{label}: {value}"
        cell.font = _FONT_META_VALUE
        cell.fill = _WASH_FILL
        cell.alignment = _ALIGN_LEFT
        # Fill merged range
        for cc in range(col_idx, end_col + 1):
            ws.cell(row=current_row, column=cc).fill = _WASH_FILL
        col_idx = end_col + 1
    # Fill remaining
    for c in range(col_idx, num_cols + 1):
        ws.cell(row=current_row, column=c).fill = _WASH_FILL
    current_row += 1

    # ── Row 5: Summary metrics (if defined) ──────────────────────────────
    if report_def.summary_fields and req.summary:
        ws.row_dimensions[current_row].height = 36
        cols_per_summary = max(1, num_cols // max(len(report_def.summary_fields), 1))
        col_idx = 1
        for sf in report_def.summary_fields:
            end_col = min(col_idx + cols_per_summary - 1, num_cols)
            if cols_per_summary > 1:
                ws.merge_cells(start_row=current_row, start_column=col_idx,
                               end_row=current_row, end_column=end_col)
            cell = ws.cell(row=current_row, column=col_idx)
            val = req.summary.get(sf["key"], "")
            formatted = format_value(val, sf.get("type", ColType.TEXT))
            cell.value = f"{sf['label']}\n{formatted}"
            cell.font = _FONT_SUMMARY_VALUE
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cc in range(col_idx, end_col + 1):
                ws.cell(row=current_row, column=cc).fill = _WHITE_FILL
            col_idx = end_col + 1
        current_row += 1

    # ── Spacer row ────────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 6
    current_row += 1

    # ── Column headers ────────────────────────────────────────────────────
    header_row = current_row
    ws.row_dimensions[current_row].height = 22
    for i, col_def in enumerate(report_def.columns):
        cell = ws.cell(row=current_row, column=i + 1)
        cell.value = col_def.header
        cell.font = _FONT_COL_HEADER
        cell.fill = _VIOLET_FILL
        cell.alignment = Alignment(
            horizontal=_col_alignment(col_def.col_type).horizontal,
            vertical="center",
        )
    current_row += 1

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(req.rows):
        ws.row_dimensions[current_row].height = 18
        is_alt = row_idx % 2 == 1
        row_fill = _WASH_FILL if is_alt else _WHITE_FILL

        for col_idx, col_def in enumerate(report_def.columns):
            cell = ws.cell(row=current_row, column=col_idx + 1)
            raw_val = row_data.get(col_def.key)

            # Write formatted string for display
            cell.value = format_value(raw_val, col_def.col_type)
            cell.font = _col_font(col_def.col_type)
            cell.fill = row_fill
            cell.alignment = _col_alignment(col_def.col_type)
            cell.border = _THIN_BORDER

        current_row += 1

    # ── Totals row ────────────────────────────────────────────────────────
    if report_def.show_totals and req.totals:
        ws.row_dimensions[current_row].height = 22
        for col_idx, col_def in enumerate(report_def.columns):
            cell = ws.cell(row=current_row, column=col_idx + 1)
            if col_idx == 0:
                label = report_def.totals_label
                if "count" in req.totals:
                    label = f"{label} ({format_value(req.totals['count'], ColType.INTEGER)} assets)"
                cell.value = label
            elif col_def.key in req.totals:
                cell.value = format_value(req.totals[col_def.key], col_def.col_type)
            else:
                cell.value = ""
            cell.font = _col_font(col_def.col_type, is_totals=True)
            cell.fill = _VIOLET_FILL
            cell.alignment = _col_alignment(col_def.col_type)
        current_row += 1

    # ── Print setup ───────────────────────────────────────────────────────
    # Repeat column header row on every printed page
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:{last_col_letter}{current_row - 1}"

    # Footer
    company_line = req.company.legal_name
    if req.company.full_address:
        company_line += f" · {req.company.full_address}"
    if req.company.registration_number:
        company_line += f" · {req.company.registration_number}"

    ws.oddFooter.left.text = company_line
    ws.oddFooter.left.size = 7
    ws.oddFooter.right.text = "Powered by VAIRIOT · Page &P of &N"
    ws.oddFooter.right.size = 7

    # ── Freeze panes below header row ─────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Output ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
